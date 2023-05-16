# import pandas as pd
# import openpyxl
# workbook = openpyxl.load_workbook("Book1.xlsx")
# df = pd.read_excel('Book1.xlsx')

# print(df)

import openpyxl
path="Book1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column
print(row,column)
for i in range(1,row):
    cell_obj = sheet_obj.cell(row = i , column = 1)
    print(cell_obj.value) 